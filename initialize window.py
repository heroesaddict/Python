import tkinter as tk
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl import load_workbook

class SampleApp(tk.Tk):
    def __init__(self):
        tk.Tk.__init__(self)
        self.entry1 = tk.Entry(self)
        self.button = tk.Button(self, text="Get", command=self.on_button)
        self.button.pack()
        self.entry1.pack()
      

    def on_button(self):
        print(self.entry1.get())
        x=self.entry1.get()
        writeCell(x);

def writeCell(x):
    
    print("pass")
    wb = Workbook()
    print("pass1")
    ws1 = wb.active
    print("pass2")
    wb = load_workbook(filename = 'C:\Python\empty_book.xlsx')
    print("pass3")
    ws = wb.worksheets[0]
    ws['A1'] = x
    wb.save('C:\Python\empty_book.xlsx')
    print("pass4")
    return
        
       

app = SampleApp()


app.mainloop()
